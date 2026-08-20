"""California Prop 65 chemical list tracker (OEHHA).

The Prop 65 listing page (oehha.ca.gov/proposition-65/proposition-65-list) is
behind an Incapsula bot challenge — a plain request gets a 200 with an empty
JS-redirect stub, no real content. The actual data file is not: OEHHA serves
the current list at a STABLE (non-dated) Excel URL that isn't gated, so this
skips the page entirely and goes straight to the file.

Unlike the ECHA fetcher (single legal document, hash the whole thing), this
source is genuinely structured — full name/CAS/mechanism/date table, ~1,000
rows. So the useful diff isn't "did the file change" but "which chemicals
were added or removed since last run." That's the actionable signal for a
pigment manufacturer: a newly listed chemical might be in the formulation.
"""

import json
import os
import time
import urllib.request
from datetime import datetime, timezone
from io import BytesIO

import openpyxl

XLSX_URL = "https://oehha.ca.gov/sites/default/files/media/downloads/proposition-65/p65chemicalslist.xlsx"

UA = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/120.0 Safari/537.36"
    )
}

STATE_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "prop65_state.json")
REPORT_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "prop65_report.json")

# Header lands here in the current workbook layout; OEHHA has moved it before
# (a change in preamble paragraph count above it would shift this). If a run
# finds zero rows or a first row that doesn't look like a chemical, the header
# row has moved and HEADER_ROW needs updating by hand.
HEADER_ROW = 12
COLUMNS = ["chemical", "type_of_toxicity", "listing_mechanism", "cas_no", "date_listed", "nsrl_madl"]


def fetch_with_retry(req, timeout, attempts=5, pause=5):
    last_err = None
    for attempt in range(attempts):
        try:
            with urllib.request.urlopen(req, timeout=timeout) as r:
                return r.read()
        except Exception as e:
            last_err = e
            if attempt < attempts - 1:
                time.sleep(pause * (2 ** attempt))
    raise RuntimeError(f"failed after {attempts} attempts: {last_err}")


def fetch_workbook(timeout=60):
    req = urllib.request.Request(XLSX_URL, headers=UA)
    data = fetch_with_retry(req, timeout)
    return openpyxl.load_workbook(BytesIO(data), read_only=True, data_only=True)


def parse_rows(wb):
    ws = wb.active
    rows = list(ws.iter_rows(min_row=HEADER_ROW + 1, values_only=True))

    header = [c.value for c in list(ws.iter_rows(min_row=HEADER_ROW, max_row=HEADER_ROW))[0]]
    if not header or "Chemical" not in str(header[0]):
        raise RuntimeError(f"row {HEADER_ROW} doesn't look like the header anymore "
                            f"(got {header[:3]!r}) — sheet layout changed, update HEADER_ROW")

    entries = {}
    for row in rows:
        chemical = row[0]
        if not chemical or not str(chemical).strip():
            continue
        record = {col: row[i] if i < len(row) else None for i, col in enumerate(COLUMNS)}
        record["chemical"] = str(record["chemical"]).strip()
        record["cas_no"] = str(record["cas_no"]).strip() if record["cas_no"] else None
        if record["date_listed"] and hasattr(record["date_listed"], "isoformat"):
            record["date_listed"] = record["date_listed"].date().isoformat()
        # CAS number is the natural unique key; a handful of multi-substance
        # listings have none, so fall back to the chemical name for those.
        key = record["cas_no"] or record["chemical"]
        entries[key] = record

    if not entries:
        raise RuntimeError("parsed 0 chemical rows — sheet layout likely changed")
    return entries


def load_state():
    if not os.path.exists(STATE_FILE):
        return {}
    with open(STATE_FILE, encoding="utf-8") as f:
        return json.load(f)


def save_state(state):
    with open(STATE_FILE, "w", encoding="utf-8") as f:
        json.dump(state, f, indent=2, sort_keys=True)


def main():
    prior = load_state()

    try:
        wb = fetch_workbook()
        current = parse_rows(wb)
    except Exception as e:
        print(f"  FAIL  Prop 65  {e}")
        with open(REPORT_FILE, "w", encoding="utf-8") as f:
            json.dump({"checked_at": datetime.now(timezone.utc).isoformat(timespec="seconds"),
                       "error": str(e)}, f, indent=2)
        raise

    added = [current[k] for k in current if k not in prior]
    removed = [prior[k] for k in prior if k not in current]

    if not prior:
        print(f"  NEW   Prop 65  baseline recorded — {len(current)} chemicals")
    else:
        print(f"  OK    Prop 65  {len(current)} chemicals, {len(added)} added, {len(removed)} removed")
        for r in added:
            print(f"    + {r['chemical']} (CAS {r['cas_no']}, listed {r['date_listed']})")
        for r in removed:
            print(f"    - {r['chemical']} (CAS {r['cas_no']})")

    save_state(current)

    with open(REPORT_FILE, "w", encoding="utf-8") as f:
        json.dump({
            "checked_at": datetime.now(timezone.utc).isoformat(timespec="seconds"),
            "total_chemicals": len(current),
            "added": added,
            "removed": removed,
        }, f, indent=2)


if __name__ == "__main__":
    main()
